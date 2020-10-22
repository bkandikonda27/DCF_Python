import pandas as pd
import finnhub
import json
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.worksheet.table import Table
import string
import FinnhubConnector


class DiscountedCashFlow:

    wb = Workbook()
    alpha = string.ascii_uppercase

    def __init__(self,ticker):
        self.ticker = ticker
        self.api = FinnhubConnector.FinnhubConnector(self.ticker)
    def change_column_width(self,column_num,ws):
        ws.column_dimensions[self.alpha[column_num-1]].width = 12
    def banner(self,val1,val2,ws, words,type):
        ws.merge_cells('{0}:{1}'.format(val1,val2))
        ws[val1].value = words
        if type == 0:
            ws[val1].fill = PatternFill('solid',fgColor='0066cc')
            ws[val1].font = Font(color='FFFFFF')
            ws[val1].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        elif type == 1:
            ws[val1].fill = PatternFill('solid', fgColor='C7DDE1')
        elif type == 2:
            ws[val1].fill = PatternFill('solid', fgColor='C7DDE1')
            ws[val1].font = Font(color='FFFFFF')
            ws[val1].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
    def assumption_cell(self, val,ws):
        ws[val].fill = PatternFill('solid',fgColor='ffff99')
    def do_banner(self,ws,ticker):
        self.banner('B5', 'R6', ws, '{0} - DCF Assumptions & Output:'.format(ticker),0)
        self.banner('B39', 'F40', ws, '{0} - FCF Projections:'.format(ticker),0)
        self.banner('G39', 'K40', ws, 'Historical', 0)
        self.banner('L39','P40',ws,'Projected',2)
    #DCF SUPPORT FUNCTION
    def percent_of_revenue(self,col,row_number,ws):
        currentcol = self.alpha[col - 1]
        activecell = ws['{0}{1}'.format(currentcol, row_number)]
        activecell.number_format = '0.00%'
        ws.cell(row=row_number, column=col,
                value='=IFERROR({0}{1}/{0}{2},{3}{4})'.format(currentcol, row_number-1, 43, self.alpha[col-2],row_number))
                # =IF({0}{1}/{0}{2} =0%,{3}{4},{0}{1}/{0}{2})
                # =IFERROR({0}{1}/{0}{2},{3}{4})
    # DCF SUPPORT FUNCTION
    def percent_Y_Y_growth(self, ws, col, row_number):
        currentcol = self.alpha[col - 1]
        pastcol = self.alpha[col - 2]
        activecell = ws['{0}{1}'.format(currentcol, row_number)]
        activecell.number_format = '0.00%'
        ws.cell(row=row_number, column=col,
                value='=IF({0}{1}/{3}{1}-1=0%,{3}{2},{0}{1}/{3}{1}-1)'.format(currentcol, row_number - 1, row_number,pastcol))
                #    =IF({0}{1}/{3}{1}-1=0%,{3}{2},{0}{1}/{3}{1}-1)
                #=IFERROR({0}{1}/{3}{1}-1,{3}{2})
    def zero_block(self,ws):
        df = pd.read_csv('datafiles/assumptions.csv')
        df = df[df['when'] == 0]
        for account in range(len(df)):
            ws.cell(row=account+8,column=2,value='{0}'.format(df.iloc[account,0]))
            cell_value = 'F{0}'.format(account+8)
            self.assumption_cell(cell_value,ws)
    def equity_value_calc(self,ws):
        ## First, Third, and Fifth Block in  Assumptions and Output Section
        df = pd.read_csv('datafiles/assumptions.csv')
        df1 = df[df['when'] == 1]

        #First Block
        for account in range(len(df1) - 2):
            ws.cell(row=account + 21, column=2, value='{0}'.format(df1.iloc[account, 0]))
            cell_value = '{0}{1}'.format(self.alpha[5],account + 21)
            self.assumption_cell(cell_value, ws)
            if account == 0:
                header = df1[df1['base'] == 'header']
                self.banner('B20','E20',ws,header.iloc[account,0],1)
                ws.cell(row=account+20,column=2,value='{0}'.format(header.iloc[account,0]))
                ws['F20'] = '=F10*F11'
                self.make_cell_accounting(ws, 6, 20)
                ender = df1[df1['base'] == 'ender']
                ws.cell(row=account + 30, column=2, value='{0}'.format(ender.iloc[account, 0]))
                self.banner('B30', 'E30', ws, ender.iloc[account, 0],1)
                ws['F30'] = '=SUM(F20:F29)'
                self.make_cell_accounting(ws,6,30)
        # Third, and Fifth Block
        for column_letter in ([8,14]):
            df2 = df[df['when'] == 3]
            for account in range(len(df2) - 1):
                if account == 0:
                    ws.cell(row=account + 19,column=column_letter,value='{0}'.format(df2.iloc[account, 0]))
                    ws['{0}19'.format(self.alpha[column_letter+3])] = '={0}17/{0}13-1'.format(self.alpha[column_letter+3])
                    ws['{0}19'.format(self.alpha[column_letter+3])].number_format = '0.00 %'
                else:
                    ws.cell(row=account + 20, column=column_letter, value='{0}'.format(df2.iloc[account, 0]))
                    ws.cell(row=account + 20, column=column_letter, value='{0}'.format(df2.iloc[account, 0]))
                    ws.cell(row=account + 20, column=column_letter + 4, value='=-F{0}'.format(account+20))
                if (column_letter == 8 and account == 0) or (column_letter == 14 and account == 0):
                    ender = df2[df2['base'] == 'ender']
                    self.banner('{0}30'.format(self.alpha[column_letter-1]), '{0}30'.format(self.alpha[column_letter+2]), ws, ender.iloc[account, 0],1)
                    ws['{0}30'.format(self.alpha[column_letter+3])] = '={0}17+SUM({0}21:{0}29)'.format(self.alpha[column_letter+3])
    def find_price_from_dcf(self,ws):
        ws = self.wb.active
        df = pd.read_csv('datafiles/assumptions.csv')
        df1 = df[df['when'] == 7]
        row_number = 33
        for account in range(len(df1)):
            for column_number in ([8,14]):
                if df1.iloc[account,0] == 'Implied Share Price from DCF:':
                    ws.cell(row=row_number,column=column_number,value=df1.iloc[account,0])
                    ws.cell(row = row_number,column=column_number+4,value='=+{0}30/F11'.format(self.alpha[column_number+3]))
                    self.make_cell_accounting(ws,column_number,row_number)
                    ws['L34'].number_format = '0.00%'
                else:
                    ws.cell(row=row_number+1,column=column_number,value=df1.iloc[account,0])
                    self.make_cell_accounting(ws, column_number, row_number)
                    ws.cell(row=row_number+1,column=column_number+4,value='=+{0}33/F11-1'.format(self.alpha[column_number+3]))
                    ws['R34'].number_format = '0.00%'

                # ws.cell(row=33, column=column_number-1, value=df1.iloc[account, 0])
                # ws.cell(row=34, column=column_number-1,value='=R33/F10-1')
    def make_cell_accounting(self,ws,col,row):
        currentcol = self.alpha[col-1]
        activecell = ws['{0}{1}'.format(currentcol, row)]
        activecell.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
    def fill_dcf_data(self,ws):
        # For extra detail like %, M...
        detail = NamedStyle(name='detail')
        detail.font = Font(size=10, italic=True)
        detail.border = Border(bottom=Side(border_style=None, color='FF000000'))

        # Get Data
        dcfnames = pd.read_csv("datafiles/dcfnames.csv")
        x = self.api.get_company_financials('ic', 'annual')
        x = pd.DataFrame(x)
        cf = self.api.get_company_financials('cf', 'annual')
        cf = pd.DataFrame(cf)
        bs = self.api.get_company_financials('bs','annual')
        year = x['year']
        self.fill_equity_value_data(ws,bs,x)

        #Tax Rate and Discount Rate
        for num in range(12,14):
            active_cell = ws['F{0}'.format(num)]
            if num == 12:
                ws['F{0}'.format(num)] = .4
            else:
                ws['F{0}'.format(num)] = .1
            active_cell.number_format = '0.00%'

        # Set years for DCF Section
        for num in range(0, 10):
            if num < 5:
                ws.cell(row=41, column=num + 7, value=year.iloc[num])
            if num >= 5:
                ws.cell(row=41, column=num + 7, value=year.iloc[4] + num - 4)

        #Set Column Width of Data to 10
        for col in range(7,22):
            self.change_column_width(col-1,ws)

        #Make DCF
        for row in range(len(dcfnames)):
            #Set up Names in Column B
            ws.cell(row=row + 42, column=2, value="{0}".format(dcfnames.iloc[row, 0]))
            #Set up type of data in each row
            if dcfnames.iloc[row,5] == "$ M" or dcfnames.iloc[row,5] == "%" or dcfnames.iloc[row,5] == "#":
                ws.cell(row=row+42,column=5,value='{0}'.format(dcfnames.iloc[row,5]))
            activecell1 = ws['E{0}'.format(row + 42)]
            activecell1.font = Font(size=10, italic=True)
            activecell1.alignment = Alignment(indent=1)
            if dcfnames.iloc[row, 4] == 'yes' or dcfnames.iloc[row, 4] == 'rev':
                activecell = ws['B{0}'.format(row + 42)]
                activecell.font = Font(size=10, italic=True)
                activecell.alignment = Alignment(indent=1)
            if dcfnames.iloc[row, 2] == 'revenue':
                for col in range(7, 12):
                    dfrev = x['revenue']
                    ws.cell(row=43, column=col, value=dfrev.iloc[col - 7])
                    self.make_cell_accounting(ws, col, 43)
                    if col != 7:
                        self.percent_Y_Y_growth(ws, col, 44)
                for col in range(12,17):
                    ws.cell(row=43, column=col, value='=IFERROR({0}{1}*(1+{2}{3}),{0}{1})'.format(self.alpha[col-2],43,self.alpha[col-1],44))
                    #    =IF({0}{1}*(1+{2}{3}) = )
                    self.make_cell_accounting(ws, col, 43)
                    cell = '{0}44'.format(self.alpha[col-1])
                    cell1 = ws[cell]
                    cell1.number_format = '0.00%'
                    cell1.value = ws['K44'].value
                    self.assumption_cell(cell,ws)
            elif dcfnames.iloc[row, 2] == 'costOfGoodsSold':
                for col in range(7,12):
                    dfrev = x['costOfGoodsSold']
                    ws.cell(row=46, column=col, value=dfrev.iloc[col - 7])
                    self.make_cell_accounting(ws, col, 46)
                    self.percent_of_revenue(col,47,ws)
                for col in range(12,17):
                    ws.cell(row=46, column=col, value='=IFERROR({0}{1}*{0}{2},{3})'.format(self.alpha[col-1],47,43,0))
                    self.make_cell_accounting(ws, col, 46)
                    cell = '{0}47'.format(self.alpha[col - 1])
                    cell1 = ws[cell]
                    cell1.number_format = '0.00%'
                    self.assumption_cell(cell, ws)
                    cell1.value = ws['K47'].value
            elif dcfnames.iloc[row, 0] == '(=) Gross Profit':
                for col in range(7, 17):
                    ws.cell(row=49, column=col,
                            value='=IFERROR({0}{1}-{0}{2},{3})'.format(self.alpha[col-1], 43, 46, '""'))
                    self.make_cell_accounting(ws, col , 49)
                    if col != 7:
                        self.percent_Y_Y_growth(ws,col,50)
            elif dcfnames.iloc[row, 2] == 'totalOperatingExpense':
                for col in range(7, 12):
                    dfrev = x['totalOperatingExpense']
                    ws.cell(row=52, column=col, value=(dfrev.iloc[col - 7]))
                    self.make_cell_accounting(ws, col, 52)
                    if col != 7:
                        self.percent_Y_Y_growth(ws,col,53)
                for col in range(12,17):
                    ws.cell(row=52, column=col, value='=IFERROR((1+{0}{1})*{2}{3},{4})'.format(self.alpha[col-1],53,self.alpha[col-2] ,52,'""'))
                    self.make_cell_accounting(ws, col, 52)

                    cell = '{0}53'.format(self.alpha[col - 1])
                    cell1 = ws[cell]
                    cell1.number_format = '0.00%'
                    cell1.value = ws['K53'].value
                    self.assumption_cell(cell, ws)
            elif dcfnames.iloc[row, 0] == '(=) Operating Income':
                for col in range(7, 17):
                    ws.cell(row=55, column=col,
                            value='=IFERROR({0}{1}-{0}{2},{3})'.format(self.alpha[col-1], 49, 52, '""'))
                    self.make_cell_accounting(ws, col, 55)
                    if col != 7:
                        self.percent_Y_Y_growth(ws,col,56)
            elif dcfnames.iloc[row, 0] == '(-) Tax on Operating Income':
                for col in range(7, 12):
                    dfrev = cf['cashTaxesPaid']
                    ws.cell(row=58, column=col, value=dfrev.iloc[col - 7])
                    self.make_cell_accounting(ws, col, 58)
                for col in range(12,17):
                    ws.cell(row=58,column=col,value='=IFERROR({0}{1}*F12,{2})'.format(self.alpha[col-1],55,0))
                    self.make_cell_accounting(ws, col, 58)
            elif dcfnames.iloc[row, 0] == '(=) NOPAT':
                for col in range(7, 17):
                    ws.cell(row=60, column=col,
                            value='=IFERROR({0}{1}-{0}{2},{3})'.format(self.alpha[col-1], 55, 58, 0))
                    self.make_cell_accounting(ws, col, 60)
                    self.percent_of_revenue(col,61,ws)
            elif dcfnames.iloc[row, 0] == '(+) Depreciation & Amortization':
                for col in range(7, 12):
                    dfrev = cf['depreciationAmortization']
                    ws.cell(row=63, column=col, value=(dfrev.iloc[col - 7]))
                    self.make_cell_accounting(ws, col, 63)
                    self.percent_of_revenue(col, 64, ws)
                for col in range(12,17):
                    ws.cell(row=63,column=col,value='=IFERROR({0}43*{0}64,{1})'.format(self.alpha[col-1],0))
                    self.make_cell_accounting(ws, col, 63)
                    cell = '{0}64'.format(self.alpha[col - 1])
                    cell1 = ws[cell]
                    cell1.number_format = '0.00%'
                    cell1.value = ws['K64'].value
                    cell_value = '{0}64'.format(self.alpha[col-1])
                    self.assumption_cell(cell_value,ws)
            elif dcfnames.iloc[row, 0] == '(+/-) Deferred Income Taxes:':
                for col in range(7, 12):
                    dfrev = bs['deferredIncomeTax']
                    ws.cell(row=66, column=col , value=dfrev.iloc[col - 7])
                    self.make_cell_accounting(ws, col , 66)
                    self.percent_of_revenue(col, 67, ws)
                for col in range(12,17):
                    ws.cell(row=66, column=col, value='=IFERROR({0}43*{0}67,{1})'.format(self.alpha[col - 1], 0))
                    self.make_cell_accounting(ws, col, 66)
                    cell = '{0}67'.format(self.alpha[col - 1])
                    cell1 = ws[cell]
                    cell1.number_format = '0.00%'
                    cell1.value = ws['K67'].value
                    cell_value = '{0}67'.format(self.alpha[col - 1])
                    self.assumption_cell(cell_value, ws)
            elif dcfnames.iloc[row, 0] == '(-) Capital Expenditure':
                for col in range(7, 12):
                    dfrev = cf['capex']
                    ws.cell(row=69, column=col , value=dfrev.iloc[col - 7])
                    self.make_cell_accounting(ws, col, 69)
                    self.percent_of_revenue(col, 70, ws)
                for col in range(12,17):
                    ws.cell(row=69, column=col, value='=IFERROR({0}43*{0}70,{1})'.format(self.alpha[col - 1], 0))
                    self.make_cell_accounting(ws, col, 69)
                    cell = '{0}70'.format(self.alpha[col - 1])
                    cell1 = ws[cell]
                    cell1.number_format = '0.00%'
                    cell1.value = ws['K70'].value
                    cell_value = '{0}70'.format(self.alpha[col - 1])
                    self.assumption_cell(cell_value, ws)
            elif dcfnames.iloc[row, 0] == '(-) Changes in Net Working Capital':
                for col in range(7, 12):
                    dfrev = cf['changesinWorkingCapital']
                    ws.cell(row=72, column=col , value=dfrev.iloc[col - 7])
                    self.make_cell_accounting(ws, col , 72)
                    self.percent_of_revenue(col, 73, ws)
                for col in range(12,17):
                    ws.cell(row=72, column=col, value='=IFERROR({0}43*{0}73,{1})'.format(self.alpha[col - 1], 0))
                    self.make_cell_accounting(ws, col, 72)
                    cell = '{0}73'.format(self.alpha[col - 1])
                    cell1 = ws[cell]
                    cell1.number_format = '0.00%'
                    cell1.value = ws['K73'].value
                    cell_value = '{0}73'.format(self.alpha[col - 1])
                    self.assumption_cell(cell_value, ws)
            elif dcfnames.iloc[row, 0] == '(=) Unlevered Free Cash Flow':
                for col in range(7, 17):
                    ws.cell(row=75, column=col,
                            value='=IFERROR({0}{1}+{0}{2}+{0}{3}+{0}{4}+{0}{5},{6})'.format(self.alpha[col-1], 72, 69,63,66,60, '""'))
                    self.make_cell_accounting(ws, col, 75)
                    if col != 7:
                        self.percent_Y_Y_growth(ws,col,76)
            elif dcfnames.iloc[row, 0] == 'Discount Period:':
                for col in range(12,17):
                    ws.cell(row=79,column=col,value=col-11)
            elif dcfnames.iloc[row, 0] == 'Discount Rate (WACC):':
                for col in range(12,17):
                    ws.cell(row=80,column=col,value='=F13')
                    cell = '{0}80'.format(self.alpha[col - 1])
                    cell1 = ws[cell]
                    cell1.number_format = '0.00%'
            elif dcfnames.iloc[row, 0] == 'Cumulative Discount Factor:':
                for col in range(12, 17):
                    row_number = 80
                    ws.cell(row=81,column=col,value='=1/((1+{0}{1})^{0}{2})'.format(self.alpha[col-1],row_number,row_number-1))
                    cell = '{0}81'.format(self.alpha[col - 1])
                    cell1 = ws[cell]
                    cell1.number_format = '0.0000'
            elif dcfnames.iloc[row, 0] == 'PV of Unlevered FCF:':
                for col in range(12,17):
                    ws.cell(row=83,column=col,value='={0}81*{0}75'.format(self.alpha[col-1]))
                    self.make_cell_accounting(ws,col,83)
                    if col !=12:
                        self.percent_Y_Y_growth(ws,col,84)
            elif dcfnames.iloc[row, 0] == 'EBITDA:':
                for col in range(7,17):
                    ws.cell(row=86,column=col,value='={0}55+{0}63'.format(self.alpha[col-1]))
                    if col != 7:
                        self.percent_Y_Y_growth(ws,col,87)
                    self.make_cell_accounting(ws,col,86)
    def multiples_method(self,ws):
        df = pd.read_csv('datafiles/assumptions.csv')
        df = df[df['when'] == 2]
        df = df[df['base'] == 'n']
        self.banner('H8', 'L8', ws, 'Terminal Value - Multiples Method:', 1)
        ws.cell(row=17, column=12, value='=L14+L13')
        self.make_cell_accounting(ws,12,17)
        self.banner('H17', 'K17', ws, 'Implied Enterprise Value:', 1)
        for account in range(len(df)):
            ws.cell(row=account+9,column=8,value=df.iloc[account,0])
            if df.iloc[account,0] == 'Median EV / EBITDA of Comps:':
                self.assumption_cell('L9',ws)
                ws['L9'].number_format = '0.00x'
                ws['L9'] = 6.5
            elif df.iloc[account,0] == 'Baseline Terminal EBITDA Multiple:':
                self.assumption_cell('L10', ws)
                ws['L10'].number_format = '0.00x'
                ws['L10'] = 6.00
            elif df.iloc[account,0] == 'Baseline Terminal Value:':
                self.make_cell_accounting(ws, 12, 11)
                ws.cell(row=account+9,column=12,value='=+L10*P86'.format())
            elif df.iloc[account,0] == 'Implied Terminal FCF Growth Rate:':
                active_cell = ws['L12']
                active_cell.number_format = '0.00%'
                ws.cell(row=account+9,column=12,value='=(L11*F13-P75)/(L11+P75)')
            elif df.iloc[account,0] == '(+) PV of Terminal Value:':
                ws.cell(row=account+9,column=12,value='=+L11*P81')
                self.make_cell_accounting(ws, 12, account+9)
            elif df.iloc[account,0] == '(+) Sum of PV of Free Cash Flows:':
                ws.cell(row=account + 9, column=12, value='=SUM($L$83:$P$83)')
                self.make_cell_accounting(ws, 12, account+9)
    def perpetuity_growth_method(self,ws):
        df = pd.read_csv('datafiles/assumptions.csv')
        df = df[df['when'] == 5]
        df = df[df['base'] == 'n']
        self.banner('N8', 'R8', ws, 'Terminal Value - Perpetuity Growth Method:', 1)
        self.banner('N17', 'Q17', ws, 'Implied Enterprise Value:', 1)
        ws.cell(row=17, column=18, value='=R14+R13')
        self.make_cell_accounting(ws,18,17)
        for account in range(len(df)):
            ws.cell(row=account + 9, column=14, value=df.iloc[account, 0])
            if df.iloc[account, 0] == 'Expected Long-Term GDP Growth:':
                self.assumption_cell('R9', ws)
                ws['R9'].number_format = '0.00%'
                ws['R9'] = 0.02
            elif df.iloc[account, 0] == 'Baseline Terminal FCF Growth Rate:':
                self.assumption_cell('R10', ws)
                ws['R10'].number_format = '0.00%'
                ws['R10'] = 0.03
            elif df.iloc[account, 0] == 'Baseline Terminal Value:':
                ws.cell(row=account + 9, column=18, value='=+P75*(1+R10)/(F13-R10)')
                self.make_cell_accounting(ws, 18, account+9)
            elif df.iloc[account, 0] == 'Implied Terminal EBITDA Multiple:':
                ws.cell(row=account + 9, column=18, value='=R11/P86')
                ws['R12'].number_format = '0.00x'
            elif df.iloc[account, 0] == '(+) PV of Terminal Value:':
                ws.cell(row=account + 9, column=18, value='=+R11*P81')
                self.make_cell_accounting(ws, 18, account + 9)
            elif df.iloc[account, 0] == '(+) Sum of PV of Free Cash Flows:':
                ws.cell(row=account + 9, column=18, value='=SUM($L$83:$P$83)')
                self.make_cell_accounting(ws,18,account+9)
    def fill_equity_value_data(self,ws, x, y):
        quote = self.api.get_stock_quote()
        ws['F9'] = self.ticker
        try:
            quote = quote['c']
            ws['F10'] = quote
        except:
            print('Error printing Quote')
        try:
            cash = x['cash']
            cashE = x['cashEquivalents']
            cashE = -cashE.iloc[4]
            cash = -cash.iloc[4]
            ws['F21'] = cash + cashE
        except:
            cashE = x['cashEquivalents']
            cashE = -cashE.iloc[4]
            ws['F21'] = cashE
        try:
            st_equity = x['cashShortTermInvestments']
            st_equity = -st_equity.iloc[4]
            ws['F22'] = st_equity
        except:
            print('Error: Equity Investments in filling Equity Data')

        try:
            total_debt = x['totalDebt']
            total_debt = total_debt.iloc[4]
            ws['F24'] = total_debt
        except:
            print('Error printing Total Debt')
        try:
            shares_outstanding = y['dilutedAverageSharesOutstanding']
            shares_outstanding = shares_outstanding.iloc[4]
            ws['F11'] = shares_outstanding
            ws['F10'] = quote
        except:
            print('Error: Printing Equity Value')










    def dcf(self):
        ws = self.wb.active
        self.do_banner(ws,self.ticker)
        self.zero_block(ws)
        self.equity_value_calc(ws)
        self.fill_dcf_data(ws)
        self.multiples_method(ws)
        self.perpetuity_growth_method(ws)
        self.find_price_from_dcf(ws)
        self.wb.save(filename="{0}{1}.xlsx".format(self.ticker,' - Bala Kandikonda'))


y = DiscountedCashFlow('MSFT')
y.dcf()

